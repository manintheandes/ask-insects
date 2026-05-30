from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
import hashlib
import json
from pathlib import Path
from typing import Callable
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from askinsects.records import EvidenceRecord, Provenance


DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID = "drosophila_suzukii_plos_climate_suitability"
SPECIES = "Drosophila suzukii"
COMMON_NAME = "spotted wing drosophila"
ARTICLE_DOI = "10.1371/journal.pone.0174318"
ARTICLE_URL = "https://journals.plos.org/plosone/article?id=10.1371/journal.pone.0174318"
LICENSE = "CC-BY-4.0"
USER_AGENT = "AskInsects/0.1 source-plane"


SUPPLEMENT_SPECS = [
    {
        "id": "s001",
        "filename": "pone.0174318.s001.docx",
        "title": "References used to compile the dataset",
        "content_type": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    },
    {
        "id": "s002",
        "filename": "pone.0174318.s002.xlsx",
        "title": "Principal components analysis for environmental-variable selection",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    },
    {
        "id": "s003",
        "filename": "pone.0174318.s003.xlsx",
        "title": "Principal-component correlations with environmental variables",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    },
    {
        "id": "s004",
        "filename": "pone.0174318.s004.xlsx",
        "title": "Moran's I spatial-autocorrelation values for selected environmental variables",
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    },
]


@dataclass(frozen=True)
class FetchBody:
    body: bytes
    content_type: str
    status: int
    final_url: str = ""


@dataclass(frozen=True)
class DrosophilaSuzukiiPlosClimateSuitabilityResult:
    source_id: str
    records: list[EvidenceRecord]
    gaps: list[dict[str, object]]
    raw_artifacts: list[str]
    requested_urls: list[str]
    file_count: int
    parsed_table_row_count: int


def _supplement_url(supplement_id: str) -> str:
    return f"https://journals.plos.org/plosone/article/file?type=supplementary&id={ARTICLE_DOI}.{supplement_id}"


def _default_fetch_body(url: str) -> FetchBody:
    request = Request(url, headers={"User-Agent": USER_AGENT, "Accept": "*/*"})
    with urlopen(request, timeout=90) as response:
        return FetchBody(
            body=response.read(),
            content_type=str(response.headers.get("content-type") or ""),
            status=int(getattr(response, "status", 200)),
            final_url=str(response.geturl()),
        )


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _safe_id(value: object) -> str:
    text = str(value or "").strip().lower()
    return "".join(ch if ch.isalnum() else "_" for ch in text).strip("_") or "unknown"


def _write_bytes(raw_dir: Path, filename: str, body: bytes) -> Path:
    raw_dir.mkdir(parents=True, exist_ok=True)
    path = raw_dir / filename
    path.write_bytes(body)
    return path


def _cell_value(value: object) -> object:
    if isinstance(value, float):
        return round(value, 8)
    return value


def _rows_from_xlsx(path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        material_rows = [
            tuple(_cell_value(value) for value in row)
            for row in sheet.iter_rows(values_only=True)
            if any(value is not None and str(value).strip() for value in row)
        ]
        if len(material_rows) < 2:
            continue
        table_title = str(material_rows[0][0] or "").strip()
        headers = [str(value or "").strip() for value in material_rows[1]]
        if headers and not headers[0]:
            headers[0] = "variable"
        for row_number, values in enumerate(material_rows[2:], start=3):
            row_payload = {
                header: _cell_value(values[index]) if index < len(values) else None
                for index, header in enumerate(headers)
                if header
            }
            if not row_payload:
                continue
            rows.append(
                {
                    "sheet": sheet.title,
                    "table_title": table_title,
                    "row_number": row_number,
                    "fields": row_payload,
                }
            )
    return rows


def _summary_record(retrieved_at: str) -> EvidenceRecord:
    return EvidenceRecord(
        record_id="swd_plos_climate_suitability:000_summary",
        lane="ecology",
        source=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
        title="Drosophila suzukii global climate-suitability model summary",
        text=(
            "PLOS ONE climate-suitability model for Drosophila suzukii used 407 occurrence sites, "
            "11 predictor variables, MaxEnt and GARP models, and 1000 bootstrap replicates. "
            "Reported model performance was AUC 0.97 for MaxEnt and 0.87 for GARP. "
            "The article identifies annual mean temperature, maximum temperature of the warmest month, "
            "mean temperature of the coldest quarter, and annual precipitation as influential MaxEnt variables."
        ),
        species=SPECIES,
        url=ARTICLE_URL,
        media_url=None,
        provenance=Provenance(
            source_id=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
            locator=f"{ARTICLE_URL}#abstract",
            retrieved_at=retrieved_at,
            license=LICENSE,
            source_url=ARTICLE_URL,
        ),
        payload={
            "atom_type": "plos_climate_model_summary",
            "doi": ARTICLE_DOI,
            "occurrence_site_count": 407,
            "model_algorithms": ["MaxEnt", "GARP"],
            "bootstrap_replicates": 1000,
            "auc": {"maxent": 0.97, "garp": 0.87},
            "influential_variables": [
                "annual mean temperature",
                "maximum temperature of the warmest month",
                "mean temperature of the coldest quarter",
                "annual precipitation",
            ],
        },
    )


def _file_manifest_record(spec: dict[str, str], response: FetchBody, raw_path: Path, retrieved_at: str) -> EvidenceRecord:
    supplement_id = spec["id"]
    return EvidenceRecord(
        record_id=f"swd_plos_climate_suitability:001_file:{supplement_id}",
        lane="ecology",
        source=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
        title=f"Drosophila suzukii PLOS climate-suitability supplement {supplement_id}: {spec['title']}",
        text=(
            f"PLOS climate-suitability supplement {supplement_id} for Drosophila suzukii. "
            f"File: {spec['filename']}. Byte size: {len(response.body)}. SHA-256: {_sha256(response.body)}."
        ),
        species=SPECIES,
        url=ARTICLE_URL,
        media_url=_supplement_url(supplement_id),
        provenance=Provenance(
            source_id=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
            locator=raw_path.as_posix(),
            retrieved_at=retrieved_at,
            license=LICENSE,
            source_url=_supplement_url(supplement_id),
        ),
        payload={
            "atom_type": "plos_climate_supplement_file",
            "supplement_id": supplement_id,
            "filename": spec["filename"],
            "title": spec["title"],
            "byte_size": len(response.body),
            "sha256": _sha256(response.body),
            "content_type": response.content_type or spec["content_type"],
            "download_url": _supplement_url(supplement_id),
            "final_url": response.final_url,
            "raw_path": raw_path.as_posix(),
        },
    )


def _table_record(supplement_id: str, raw_path: Path, row: dict[str, object], retrieved_at: str) -> EvidenceRecord:
    fields = row["fields"] if isinstance(row.get("fields"), dict) else {}
    row_number = int(row.get("row_number") or 0)
    first_value = next(iter(fields.values()), "")
    if supplement_id == "s002":
        atom_type = "plos_climate_pca_row"
        label = f"principal component {fields.get('PC', first_value)}"
    elif supplement_id == "s003":
        atom_type = "plos_climate_variable_correlation_row"
        label = f"environmental variable {first_value}"
    else:
        atom_type = "plos_climate_moran_i_row"
        label = f"environmental variable {first_value}"
    return EvidenceRecord(
        record_id=f"swd_plos_climate_suitability:010_table:{supplement_id}:row:{row_number:04d}:{_safe_id(first_value)}",
        lane="ecology",
        source=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
        title=f"Drosophila suzukii climate-suitability {supplement_id} row {row_number}: {label}",
        text=(
            f"PLOS Drosophila suzukii climate-suitability supplementary table row for {label}. "
            f"Fields: {json.dumps(fields, sort_keys=True)}."
        ),
        species=SPECIES,
        url=ARTICLE_URL,
        media_url=None,
        provenance=Provenance(
            source_id=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
            locator=f"{raw_path.as_posix()}#sheet/{row.get('sheet')}/row/{row_number}",
            retrieved_at=retrieved_at,
            license=LICENSE,
            source_url=_supplement_url(supplement_id),
        ),
        payload={
            "atom_type": atom_type,
            "supplement_id": supplement_id,
            "sheet": row.get("sheet"),
            "table_title": row.get("table_title"),
            "row_number": row_number,
            "fields": fields,
        },
    )


def _raster_gap_record(retrieved_at: str) -> EvidenceRecord:
    return EvidenceRecord(
        record_id="swd_plos_climate_suitability:900_gap:suitability_raster_files_not_downloadable",
        lane="ecology",
        source=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
        title="Drosophila suzukii climate-suitability gap: raster suitability grids not downloadable",
        text=(
            "The PLOS climate-suitability article exposes supplementary references and model-variable tables, "
            "but not raw MaxEnt/GARP raster suitability grids. Ask Insects stores this as a source gap instead "
            "of pretending figure-map images are queryable raster files."
        ),
        species=SPECIES,
        url=ARTICLE_URL,
        media_url=None,
        provenance=Provenance(
            source_id=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
            locator=f"{ARTICLE_URL}#supporting-information",
            retrieved_at=retrieved_at,
            license=LICENSE,
            source_url=ARTICLE_URL,
        ),
        payload={
            "atom_type": "source_gap",
            "reason": "plos_suitability_raster_files_not_downloadable",
            "available_supplements": [spec["id"] for spec in SUPPLEMENT_SPECS],
        },
    )


def fetch_drosophila_suzukii_plos_climate_suitability_records(
    *,
    raw_dir: Path,
    fetch_body: Callable[[str], FetchBody] | None = None,
    retrieved_at: str | None = None,
) -> DrosophilaSuzukiiPlosClimateSuitabilityResult:
    fetch = fetch_body or _default_fetch_body
    retrieved = retrieved_at or datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    records = [_summary_record(retrieved), _raster_gap_record(retrieved)]
    gaps: list[dict[str, object]] = []
    raw_artifacts: list[str] = []
    requested_urls: list[str] = []
    parsed_table_row_count = 0

    for spec in SUPPLEMENT_SPECS:
        supplement_id = str(spec["id"])
        url = _supplement_url(supplement_id)
        requested_urls.append(url)
        try:
            response = fetch(url)
            if response.status >= 400 or not response.body:
                raise RuntimeError(f"HTTP {response.status}")
            raw_path = _write_bytes(raw_dir, str(spec["filename"]), response.body)
            raw_artifacts.append(raw_path.as_posix())
            records.append(_file_manifest_record(spec, response, raw_path, retrieved))
            if str(spec["filename"]).endswith(".xlsx"):
                table_rows = _rows_from_xlsx(raw_path)
                for row in table_rows:
                    records.append(_table_record(supplement_id, raw_path, row, retrieved))
                parsed_table_row_count += len(table_rows)
        except Exception as exc:
            gaps.append(
                {
                    "source": DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
                    "reason": "plos_climate_supplement_fetch_or_parse_failed",
                    "supplement_id": supplement_id,
                    "url": url,
                    "error": str(exc),
                    "retrieved_at": retrieved,
                }
            )

    return DrosophilaSuzukiiPlosClimateSuitabilityResult(
        source_id=DROSOPHILA_SUZUKII_PLOS_CLIMATE_SUITABILITY_SOURCE_ID,
        records=records,
        gaps=gaps,
        raw_artifacts=raw_artifacts,
        requested_urls=requested_urls,
        file_count=len(raw_artifacts),
        parsed_table_row_count=parsed_table_row_count,
    )
